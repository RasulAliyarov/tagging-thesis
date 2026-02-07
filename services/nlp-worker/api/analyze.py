from ast import List
from typing import Optional
from urllib import request
from fastapi import APIRouter, Depends, HTTPException
from core.deps import get_current_user, get_db 
from google import genai
import uuid, os, json, datetime
from core.config import settings
from pydantic import BaseModel
from bson import ObjectId
from fastapi.encoders import jsonable_encoder
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font, Alignment


router = APIRouter(prefix="/api/analyze", tags=["Analysis"])
client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

class AnalyzeRequest(BaseModel):
    text: str

@router.post("/")
async def analyze_text(
    request: AnalyzeRequest, 
    current_user: dict = Depends(get_current_user), 
    db = Depends(get_db)
):
    try:
        # 1. Get the incremental number scoped to THIS user
        last_record = await db.analyses.find_one(
            {"owner": current_user["username"]}, 
            sort=[("num", -1)]
        )
        next_num = (last_record["num"] + 1) if last_record else 1

        # 2. Prepare the prompt
        prompt = (
            f"Analyze the following text: '{request.text}'. "
            f"Provide analysis in strict JSON format with these keys: "
            f"sentiment (Positive/Neutral/Negative), priority (Low/Medium/High), "
            f"confidence (0.0-1.0), and tags: exactly 3 lowercase tags "
            f"1) text type 2) main issue 3) specific context"
        )

        # 3. Generating content using Gemini
        response = client.models.generate_content(
            model=settings.GEMINI_MODEL,
            contents=prompt
        )
        
        # 4. Manual JSON Cleaning and Parsing (Your proven method)
        raw_text = response.text.replace("```json", "").replace("```", "").strip()
        ai_data = json.loads(raw_text)
        
        # 5. Construct the final record for MongoDB
        analysis_record = {
            "id": str(uuid.uuid4()),
            "owner": current_user["username"], # Link to current user
            "num": next_num,
            "text": request.text,
            "sentiment": ai_data.get("sentiment", "Neutral"),
            "priority": ai_data.get("priority", "Medium"),
            "confidence": float(ai_data.get("confidence", 0.0)),
            "tags": ai_data.get("tags", []),
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat()
        }
        
        # 6. Persist data to MongoDB
        await db.analyses.insert_one(analysis_record.copy())
        
        return analysis_record

    except Exception as e:
        print(f"Analysis Failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class AnalysisResult(BaseModel):
    id: str
    owner: str
    num: int
    text: str
    timestamp: str
    sentiment: str
    priority: str
    confidence: float
    tags: list[str]
    
@router.get("/history", response_model=list[AnalysisResult])
async def get_my_analyses(current_user: dict = Depends(get_current_user), db = Depends(get_db)):
    cursor = db.analyses.find({"owner": current_user["username"]}).sort("timestamp", -1)
    return await cursor.to_list(length=100)


@router.delete("/history/{item_id}")
async def delete_analysis(
    item_id: str, 
    current_user: dict = Depends(get_current_user), 
    db = Depends(get_db)
):
    """Deletes a record by the custom UUID 'id' field"""
    result = await db.analyses.delete_one({"id": item_id, "owner": current_user["username"]})
    if result.deleted_count == 1:
        return {"status": "success", "message": "Record deleted"}
    raise HTTPException(status_code=404, detail="Record not found")

class AnalysisUpdate(BaseModel):
    text: Optional[str] = None
    sentiment: Optional[str] = None
    priority: Optional[str] = None
    tags: Optional[list[str]] = None
    
    
@router.put("/history/{item_id}", response_model=AnalysisResult)
async def update_analysis(
    item_id: str, 
    data: AnalysisUpdate, 
    current_user: dict = Depends(get_current_user), 
    db = Depends(get_db)
):
    """Updates a record by the custom UUID 'id' field, allowing partial updates"""
    # filter out None values to allow partial updates
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid data fields provided")

    # search and update the record, ensuring it belongs to the current user
    result = await db.analyses.find_one_and_update(
        {"id": item_id, "owner": current_user["username"]},
        {"$set": update_data},
        return_document=True
    )

    if result:
        return result
    raise HTTPException(status_code=404, detail="Record not found or access denied")


import pandas as pd
import asyncio
from fastapi import UploadFile, File

@router.post("/batch")
async def batch_analyze(
    file: UploadFile = File(...), 
    current_user: dict = Depends(get_current_user), 
    db = Depends(get_db)
):
    """Endpoint to handle batch analysis from an uploaded CSV file. The CSV should have a single column with text data."""
    try:
        # read CSV into DataFrame
        df = pd.read_csv(file.file)
        if df.empty:
            raise ValueError("CSV is empty")
        target_column = df.columns[0]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")
    
    # get the next incremental number for this user
    last_record = await db.analyses.find_one(
        {"owner": current_user["username"]}, 
        sort=[("num", -1)]
    )
    current_num = (last_record["num"] + 1) if last_record else 1
    
    final_results = []
    
    for _, row in df.iterrows():
        raw_text = str(row[target_column]).strip()
        if not raw_text: continue
        
        try:
            prompt = (
            f"Analyze the following text: '{raw_text}'. "
            f"Provide analysis in strict JSON format with these keys: "
            f"sentiment (Positive/Neutral/Negative), priority (Low/Medium/High), "
            f"confidence (0.0-1.0), and tags: exactly 3 lowercase tags "
            f"1) text type 2) main issue 3) specific context"
        )
            response = client.models.generate_content(
                model=settings.GEMINI_MODEL, 
                contents=prompt
            )
            
            ai_text = response.text.replace("```json", "").replace("```", "").strip()
            ai_data = json.loads(ai_text)
            
            record = {
                "id": str(uuid.uuid4()),
                "owner": current_user["username"], # Привязка к юзеру
                "num": current_num,
                "text": raw_text,
                "sentiment": ai_data.get('sentiment', 'Neutral'),
                "priority": ai_data.get('priority', 'Low'),
                "confidence": float(ai_data.get('confidence', 0.5)),
                "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                "tags": ai_data.get('tags', [])
            }
            
            await db.analyses.insert_one(record)
            final_results.append(record)
            current_num += 1
            
            # Rate limiting for Gemini API
            await asyncio.sleep(2) 
            
        except Exception as e:
            print(f"Skipping row due to error: {e}")
            continue
            
    return final_results


# Excel export endpoint
@router.get("/export/excel")
async def export_to_excel(
    current_user: dict = Depends(get_current_user),
    db = Depends(get_db)
):
    cursor = db.analyses.find({"owner": current_user["username"]}).sort("num", 1)
    data = await cursor.to_list(length=1000)
    
    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    df = pd.DataFrame(data)
    cols_to_drop = ['_id', 'owner', 'id'] # Убираем лишнее
    df = df.drop(columns=[col for col in cols_to_drop if col in df.columns])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Analysis')
        
        workbook = writer.book
        worksheet = writer.sheets['Analysis']

        # Стили для заголовков
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Стили для тональности
        fills = {
            "Positive": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Green
            "Negative": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # Red
            "Neutral": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Gray
        }

        # Применяем стили
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

            # Если это колонка sentiment, красим ячейки
            if column.lower() == 'sentiment':
                for row_idx in range(2, len(df) + 2):
                    val = worksheet.cell(row=row_idx, column=col_idx).value
                    if val in fills:
                        worksheet.cell(row=row_idx, column=col_idx).fill = fills[val]

        # Автоподбор ширины колонок
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="report.xlsx"'}
    )